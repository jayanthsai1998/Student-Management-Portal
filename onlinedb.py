import sys
import os
import MySQLdb as db
import click
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import fonts
from openpyxl.styles import Font
from copy import copy

# import django
# print(os.getcwd().split("\\")[-1])
# os.environ['DJANGO_SETTINGS_MODULE'] = os.getcwd().split("\\")[-1] + ".settings"
# django.setup()

from onlineproject import settings
#from onlineapp.models import *


conn = db.connect ("127.0.0.1", "root", "Tom&0erry")
cur = conn.cursor()
@click.group()

def dbsql():
    pass

@dbsql.command()
def cleardata():
    """Clears the data in the table"""
    db_name = settings.DATABASES['default']['NAME']
    conn.select_db(db_name)
    table_names = ["onlineapp_college","onlineapp_student","onlineapp_mocktest1"]

    for table in table_names:
        sql = "DELETE FROM %s"%(table)
        cur.execute(sql)
    conn.commit()

@dbsql.command()
def createdb():
    """creates the db in django settings using the creds specified there"""
    create_db_name = settings.DATABASES['default']['NAME']

    click.echo("Database to be created " + create_db_name)
    conn = db.connect("127.0.0.1", "root", "Tom&0erry")
    cur = conn.cursor()

    # To begin with a clean slate.
    cur.execute("DROP DATABASE IF EXISTS " + create_db_name)
    sql = 'CREATE DATABASE ' + create_db_name
    cur.execute(sql)
    conn.commit()
    click.echo("DB successfully created..!!!")


@dbsql.command()
def dropdb():
    """drops the db in django settings using the creds specified there"""

    drop_db_name = settings.DATABASES['default']['NAME']
    click.echo("Database name to be dropped is: " + drop_db_name)
    conn = db.connect("127.0.0.1", "root", "Tom&0erry")
    conn.select_db(drop_db_name)
    cur = conn.cursor()
    cur.execute("DROP DATABASE IF EXISTS " + drop_db_name)
    conn.commit()
    click.echo("DB dropped successfully..!!!")


@dbsql.command()
@click.argument('files', nargs = 2)
def populatedb(files):
    '''Imports the data from the files passed as arguments to tables'''
    os.chdir("C:\\work\\appscourse\\djangoproject\\onlineproject")
    students_file = files[0]
    marks_file = files[1]

    students_wb = load_workbook(students_file)
    marks_wb = load_workbook(marks_file)

    fp = 1
    college_data = []
    clg_order = [0,1,3,2,4]
    clg_sheet = students_wb["Colleges"]
    for i in range(2, clg_sheet.max_row + 1):
        cs = []
        for j in clg_order:
            if(j==0):
                cs.append(fp)
                fp += 1
            else:
                cs.append(clg_sheet.cell(row = i,column = j).value)
        college_data.append(cs)

    conn = db.connect("127.0.0.1", "root", "Tom&0erry")
    db_name = settings.DATABASES['default']['NAME']
    conn.select_db(db_name)
    cur = conn.cursor()

    for i in range(len(college_data)):
        try:
            sql = """INSERT INTO onlineapp_college VALUES """+ str(tuple(college_data[i]))
            cur.execute(sql)
        except db.Error as e:
            print("Error %d: %s" % (e.args[0], e.args[1]))
            sys.exit(1)

    conn.commit()

    fp = 1
    students_data_1 = []
    std_order = [0, 1, 6, 3, 4, 5, 2]
    stud_sheet = students_wb["Current"]
    for i in range(2, stud_sheet.max_row + 1):
        ls = []
        for j in std_order:
            if(j == 0):
                ls.append(fp)
                fp +=1
            elif(j==6):
                ls.append("NULL")
            elif(j==5):
                ls.append("False")
            elif(j==2):
                name = stud_sheet.cell(row = i, column = j).value
                for k in college_data:
                    if(k[3] == name):
                        ls.append(k[0])
                        break
            else:
                ls.append(stud_sheet.cell(row = i, column = j).value)
        students_data_1.append(ls)
    conn.commit()
    conn = db.connect("127.0.0.1", "root", "Tom&0erry")
    db_name = settings.DATABASES['default']['NAME']
    conn.select_db(db_name)
    cur = conn.cursor()

    for i in range(len(students_data_1)):
        try:
            sql = "INSERT INTO onlineapp_student values(%d,'%s',NULL,'%s','%s',False,%d)" %(students_data_1[i][0],students_data_1[i][1],students_data_1[i][3],students_data_1[i][4],students_data_1[i][6])
            #print(students_data_1[i][4])
            #print(sql)
            cur.execute(sql)
        except db.Error as e:
            print("Error %d: %s" % (e.args[0], e.args[1]))
            sys.exit(1)
    conn.commit()
    students_data_2 = []
    stud_sheet = students_wb["Deletions"]

    for i in range(2, stud_sheet.max_row + 1):
        ls = []
        for j in std_order:
            if(j==0):
                ls.append(fp)
                fp += 1
            elif(j==6):
                ls.append("NULL")
            elif(j==5):
                ls.append("True")
            elif(j == 2):
                name = stud_sheet.cell(row=i, column=j).value
                for k in college_data:
                    if(k[3] == name):
                        ls.append(k[0])
                        break
            else:
                ls.append(stud_sheet.cell(row=i, column=j).value)
        students_data_2.append(ls)



    conn = db.connect("127.0.0.1", "root", "Tom&0erry")
    db_name = settings.DATABASES['default']['NAME']
    conn.select_db(db_name)
    cur = conn.cursor()



    for i in range(len(students_data_2)):
        try:

            sql = """INSERT INTO onlineapp_student values(%d,'%s',NULL,'%s','%s',True,%d)""" %(students_data_2[i][0],students_data_2[i][1],students_data_2[i][3],students_data_2[i][4],students_data_2[i][6])
            #print(students_data_1[i][4])
            #c=Student(name=students_data[i][1],email = students_data[i][2],db_folder=students_data[i][4],college_id=College.objects.get(acronym=students_data[i][6]))
            #c.save()
            cur.execute(sql)
        except db.Error as e:
            print("Error %d: %s" % (e.args[0], e.args[1]))
            sys.exit(1)

    conn.commit()

    fp = 1
    mark_order=[0,2,3,4,5,6,1]
    marks_data = []
    mark_sheet = marks_wb.active
    for i in range(2, mark_sheet.max_row + 1):
        ls = []
        for j in mark_order:
            if(j == 0):
                ls.append(fp)
                fp += 1
            elif(j == 1):
                name = mark_sheet.cell(row = i, column = j).value.split('_')[2]
                flag=0
                for k in students_data_1:
                    if(k[4].lower() == name):
                        ls.append(k[0])
                        flag = 1
                        break
                if(flag == 0):
                    for v in students_data_2:
                        if(v[4].lower() == name):
                            ls.append(v[0])
                            break
            else:
                ls.append(int(mark_sheet.cell(row = i, column = j).value))
        #print(ls)
        marks_data.append(ls)

    conn = db.connect("127.0.0.1", "root", "Tom&0erry")
    db_name = settings.DATABASES['default']['NAME']
    conn.select_db(db_name)
    cur = conn.cursor()

    for i in range(len(marks_data)):
        try:
            # c=MockTest1(problem1 = marks_data[i][1],problem2 = marks_data[i][2],problem3 = marks_data[i][3],problem4 = marks_data[i][4],total = marks_data[i][5], student_id=Student.objects.get(name=marks_data[i][6]))
            # c.save()
            #sql = """INSERT INTO onlineapp_mocktest1 VALUES"""+str(tuple(marks_data[i]))
            sql = """INSERT INTO onlineapp_mocktest1 VALUES(%d,%d,%d,%d,%d,%d,%d)""" %(marks_data[i][0],marks_data[i][1],marks_data[i][2],marks_data[i][3],marks_data[i][4],marks_data[i][5],marks_data[i][6])
            cur.execute(sql)
        except db.Error as e:
            print("Error %d: %s" % (e.args[0], e.args[1]))
            sys.exit(1)
    conn.commit()


    # cur = conn.cursor()
    # cur.execute("SELECT * FROM STUDENTS")
    # print(cur.fetchall())
    #
    # cur.execute("SELECT * FROM MARKS")
    # print(cur.fetchall())


conn.close ()


if __name__ == "__main__":
    dbsql()